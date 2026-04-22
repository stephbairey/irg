"""
Song Archive Consolidation Script
Parses WP XML exports from main song site and Seattle site,
normalizes, extracts metadata, detects duplicates, and outputs
consolidated JSON + review spreadsheet + stats.
"""

import json
import re
import xml.etree.ElementTree as ET
from html import unescape
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = Path(__file__).resolve().parent.parent
SOURCES_DIR = BASE_DIR / "sources"

NS = {
    "wp": "http://wordpress.org/export/1.2/",
    "content": "http://purl.org/rss/1.0/modules/content/",
    "dc": "http://purl.org/dc/elements/1.1/",
    "excerpt": "http://wordpress.org/export/1.2/excerpt/",
}

CREATOR_TO_GAGGLE = {
    "Kingston": "Kingston",
    "Tucson": "Tucson",
    "Tuscon3": "Tucson",
    "Triangle NC": "Triangle",
    "Triangle NC 2": "Triangle",
    "NYC Metro": "New York City Metro",
    "NewMexico2": "New Mexico",
    "New Mexico": "New Mexico",
    "Peninsula": "San Francisco Bay Area",
    "Davis": "Davis",
    "Piedmont": "Piedmont",
    "Piedmont2": "Piedmont",
    "Victoria": "Victoria",
    "Seattle": "Seattle",
    "Madison3": "Madison",
    "Madison 2": "Madison",
    "Tuolumne1": "Tuolumne & Calaveras Counties",
    "Montreal": "Montreal",
    "Montreal2": "Montreal",
    "Detroit": "Metro Detroit",
    "Mariposa Mobile": "Tompkins County",
    "Calgary": "Calgary",
    "Westerly": "Greater Westerly RI",
    "Fresno": "Fresno",
    "Olympia": "Olympia",
    "Sacramento1": "Sacramento",
    "Green Bay": "Green Bay",
    "Grant County 1": "Grant County",
    "SoFlo": "South Florida",
    "Santa Cruz": "Santa Cruz",
    "Central Florida": "Central Florida",
    "Milwaukee": "Milwaukee",
    "Rossmoor1": "Rossmoor/Walnut Creek",
    "Westchester": "Westchester",
    "Grays Harbor": "Gray's Harbor",
    "BowlingGreen1": "Bowling Green",
    "JillH": "Kentuckiana",
    "Portland": "Portland",
    "San Jose": "San Jose & Santa Clara County",
    "Pittsburgh1": "Pittsburgh",
    "Toronto3": "Toronto",
    "Winnipeg": "Winnipeg",
    "El Dorado West Slope": "El Dorado County West Slope",
    "Mendocino": "Mendocino",
    "CathyT": "Charlotte",
    "None": "Unknown",
}

WEB_GRANNY_CREATORS = {"ragin8", "C0pdb"}

# Normalize free-text gaggle postmeta values to canonical names.
# Multi-gaggle entries get a canonical "Multiple" prefix with details preserved.
GAGGLE_POSTMETA_NORMALIZE = {
    "Barrie": "Barrie",
    "Boston": "Boston",
    "Boston, Mass (Sue) and Triangle NC (Vicki)": "Multiple (Boston, Triangle)",
    "Bowling Green Raging Grannies": "Bowling Green",
    "Boxton": "Boston",
    "Calgary": "Calgary",
    "Calgary Raging Grannies": "Calgary",
    "Calgary, Alberta Canada": "Calgary",
    "Central Florida Raging Grannies": "Central Florida",
    "Charlotte": "Charlotte",
    "Davis CA": "Davis",
    "Davis, California": "Davis",
    "Detriot": "Metro Detroit",
    "Detroit": "Metro Detroit",
    "Detroit/NYC Metro": "Multiple (Metro Detroit, New York City Metro)",
    "Fresno": "Fresno",
    "Fresno / Redding / Rochester / Rochester": "Multiple (Fresno, Rochester)",
    "Grant County Raging Grannies": "Grant County",
    "Grays Harbor": "Gray's Harbor",
    "Greater Westerly": "Greater Westerly RI",
    "Greater Westerly Grannies": "Greater Westerly RI",
    "Green Bay": "Green Bay",
    "Green Bay Raging Grannies": "Green Bay",
    "Greys Harbor/Rochester": "Multiple (Gray's Harbor, Rochester)",
    "Halifax": "Halifax",
    "Halton": "Halton",
    "Kentuckiana Raging Grannies": "Kentuckiana",
    "Kingston": "Kingston",
    "Kingston / Tucson / North Carolina": "Multiple (Kingston, Tucson, Triangle)",
    "Madison": "Madison",
    "Madison WI and thereabouts": "Madison",
    "Mariposa Mobile": "Tompkins County",
    "Mendocino": "Mendocino",
    "Milwaukee": "Milwaukee",
    "Montreal": "Montreal",
    "Montreal Raging Grannies": "Montreal",
    "Montreal?": "Montreal",
    "Montréal Raging Grannies": "Montreal",
    "NC": "Triangle",
    "NYC Metro": "New York City Metro",
    "NYC Metro/WOWW": "Multiple (New York City Metro, WOWW)",
    "NYC MetroRochester": "Multiple (New York City Metro, Rochester)",
    "NYC MetroWOWW": "Multiple (New York City Metro, WOWW)",
    "Nancy (Detroit) and Vicki (Triangle NC)": "Multiple (Metro Detroit, Triangle)",
    "New Jersey": "New Jersey",
    "New Mexico": "New Mexico",
    "New Mexico Raging Grannies": "New Mexico",
    "North Carolina": "Triangle",
    "Not listed": "Unknown",
    "Olympia Raging Grannies": "Olympia",
    "Original Lyrics: Triangle. Modifications: El Dorado County West Slope": "Multiple (Triangle, El Dorado County West Slope)",
    "Original: Probably San Francisco Bay gaggle. This version: Tucson gaggle.": "Multiple (San Francisco Bay Area, Tucson)",
    "Originally Fairgrove/San Jose (Revision by Tucson)": "Multiple (San Jose & Santa Clara County, Tucson)",
    "Originally Triangle Gaggle (updated version: Tucson Raging Grannies)": "Multiple (Triangle, Tucson)",
    "Ottawa": "Ottawa",
    "Ottawa Raging Grannies, Montreal Raging Grannies": "Multiple (Ottawa, Montreal)",
    "Peterborough": "Peterborough",
    "Piedmont Raging Grannies": "Piedmont",
    "Pittsburgh": "Pittsburgh",
    "Pittsburgh Gaggle": "Pittsburgh",
    "Portland": "Portland",
    "Portland (OR)": "Portland",
    "Raging Grannies Action League": "San Francisco Bay Area",
    "Raging Grannies Action League (SF Bay Area)": "San Francisco Bay Area",
    "Raging Grannies Action League (San Francisco Bay Area)": "San Francisco Bay Area",
    "Raging Grannies SF Bay Area/Palo Alto": "San Francisco Bay Area",
    "Raging Grannies of Madison & Dane County WI": "Madison",
    "Raging Grannies of Madison WI": "Madison",
    "Raging Grannies of Olympia": "Olympia",
    "Raging Grannies of Olympia, WA": "Olympia",
    "Raging Grannies of Tuolumne & Calaveras": "Tuolumne & Calaveras Counties",
    "Raging Grannies, Guelph": "Guelph",
    "Regina": "Regina",
    "Rochester": "Rochester",
    "Rochester (NY)": "Rochester",
    "Rochester / North Carolina": "Multiple (Rochester, Triangle)",
    "Rochester, NY": "Rochester",
    "Rochester/SoFlo/North Carolina": "Multiple (Rochester, South Florida, Triangle)",
    "RochesterSoFlo": "Multiple (Rochester, South Florida)",
    "RochesterSoFloNorth Carolina": "Multiple (Rochester, South Florida, Triangle)",
    "RochesterSouth Florida": "Multiple (Rochester, South Florida)",
    "Rossmoor": "Rossmoor/Walnut Creek",
    "SF Bay Area": "San Francisco Bay Area",
    "SF Bay Area (SF Peninsula)": "San Francisco Bay Area",
    "SF Bay Area Raging Grannies": "San Francisco Bay Area",
    "SF Bay Area/SF Peninsula": "San Francisco Bay Area",
    "SF Bay Area/SF Peninsula Raging Grannies": "San Francisco Bay Area",
    "Sacramento Raging Grannies": "Sacramento",
    "San Francisco Bay Area": "San Francisco Bay Area",
    "San Francisco Peninsula": "San Francisco Bay Area",
    "San Jose": "San Jose & Santa Clara County",
    "Santa Cruz": "Santa Cruz",
    "Santa CruzSoFloRochester": "Multiple (Santa Cruz, South Florida, Rochester)",
    "Santa Fe": "Santa Fe",
    "Santa Fe Raging Grannies": "Santa Fe",
    "Seattle": "Seattle",
    "Simcoe-Grey": "Simcoe-Grey",
    "SoFlo": "South Florida",
    "Somewhere in California": "Unknown (California)",
    "South Florida": "South Florida",
    "St John": "St. John",
    "St Paul": "St. Paul",
    "The Raging Grannies of Madison": "Madison",
    "Toronto": "Toronto",
    "Triangle (NC)": "Triangle",
    "Triangle (NC) and Central Florida": "Multiple (Triangle, Central Florida)",
    "Triangle (North Carolina)": "Triangle",
    "Triangle NC": "Triangle",
    "Triangle Raging Grannies": "Triangle",
    "Tucson": "Tucson",
    "Tucson / North Carolina": "Multiple (Tucson, Triangle)",
    "Tucson Raging Grannies": "Tucson",
    "Tucson, AZ": "Tucson",
    "Tuscon": "Tucson",
    "Unknown": "Unknown",
    "Vancouver": "Vancouver",
    "Vancouver, Montreal": "Multiple (Vancouver, Montreal)",
    "Vancouver/Rochester": "Multiple (Vancouver, Rochester)",
    "Various": "Multiple (Various)",
    "Vicki: Triangle (North Carolina) Rosalia: Detroit-Windsor": "Multiple (Triangle, Metro Detroit)",
    "Victoria": "Victoria",
    "Victoria, Seattle, Rochester": "Multiple (Victoria, Seattle, Rochester)",
    "WOWW": "WOWW",
    "Western Mass": "Western Massachusetts",
    "Westchester": "Westchester",
    "Winnipeg": "Winnipeg",
    "Wolfville": "Wolfville",
    "Wolfville Nova Scotia Raging Grannies": "Wolfville",
    "Woodstock": "Woodstock",
}

MAIN_CATEGORIES_VALID = {
    "Business & Economy",
    "Education",
    "Environment & Energy",
    "Government & Politics",
    "Health Care/Healthcare",
    "Holiday & Celebrations",
    "Human & Civil Rights",
    "Labor & Worker Rights",
    "Local Issues",
    "Soldiers & Veterans",
    "War & Peace",
    "Women's Issues",
    "World Issues",
}

MAIN_CATEGORIES_DISCARD = {"About Us", "Uncategorized"}

PRESERVE_TAGS = {"strong", "b", "em", "i", "u"}


def get_meta(item, key):
    for m in item.findall("wp:postmeta", NS):
        k = m.find("wp:meta_key", NS).text
        if k == key:
            val = m.find("wp:meta_value", NS).text
            if val and val.strip() and val.strip() != "None":
                return val.strip()
    return ""


def get_categories(item):
    cats = []
    for c in item.findall("category"):
        if c.get("domain") == "category":
            text = unescape(c.text or "")
            if text:
                cats.append(text)
    return cats


def get_tags(item):
    tags = []
    for c in item.findall("category"):
        if c.get("domain") == "post_tag":
            text = unescape(c.text or "")
            if text:
                tags.append(text)
    return tags


def extract_metadata_from_content(raw_html):
    """Extract tune, songwriter, gaggle from content body.
    Returns (tune, songwriter, gaggle, source_notes, cleaned_lines_to_remove)."""
    tune = ""
    songwriter = ""
    gaggle = ""
    source_notes = ""
    lines_to_remove = []

    search_region = raw_html[:1500]

    # Tune patterns
    tune_patterns = [
        r'(?:tune|sung to(?: the tune of)?|melody)\s*[:=]\s*["\u201c]?(.+?)["\u201d]?\s*(?:\)|$|<)',
        r'\((?:tune|sung to(?: the tune of)?|melody)\s*[:=]\s*["\u201c]?(.+?)["\u201d]?\s*\)',
        r'<em>\s*\(?(?:tune|sung to(?: the tune of)?|melody)\s*[:=]\s*(.+?)\s*\)?\s*</em>',
        r'<li>\s*<em>\s*\(?(?:tune|sung to(?: the tune of)?|melody)\s*[:=]?\s*(.+?)\s*\)?\s*</em>\s*</li>',
    ]
    for pat in tune_patterns:
        m = re.search(pat, search_region, re.IGNORECASE)
        if m:
            t = m.group(1) if m.lastindex and m.lastindex >= 1 else m.group(1)
            t = re.sub(r"<[^>]+>", "", t).strip().rstrip(")")
            if t and len(t) < 200:
                tune = t
                lines_to_remove.append(m.group(0))
                break

    # Songwriter patterns
    songwriter_patterns = [
        r'<em>\s*(?:by|written by|lyrics by|words by)\s+(.+?)\s*</em>',
        r'<li>\s*<em>\s*(?:by|written by|lyrics by|words by)\s+(.+?)\s*</em>\s*</li>',
        r'(?:^|\n)\s*(?:by|written by|lyrics by|words by)\s+([A-Z][a-zA-Z\s\-\.\']+)',
    ]
    for pat in songwriter_patterns:
        m = re.search(pat, search_region, re.IGNORECASE)
        if m:
            s = re.sub(r"<[^>]+>", "", m.group(1)).strip()
            if s and len(s) < 100 and not any(
                w in s.lower() for w in ["verse", "chorus", "stanza"]
            ):
                songwriter = s
                lines_to_remove.append(m.group(0))
                break

    # Gaggle from content
    gaggle_patterns = [
        r"from(?:\s+the)?\s+(.+?)\s+(?:Raging\s+)?Grannies",
        r"(?:Raging\s+)?Grannies\s+(?:of|from)\s+(.+?)(?:\s+Songbook|\s*$|\s*<)",
    ]
    for pat in gaggle_patterns:
        m = re.search(pat, search_region, re.IGNORECASE)
        if m:
            g = re.sub(r"<[^>]+>", "", m.group(1)).strip()
            if g and len(g) < 80:
                gaggle = g
                break

    return tune, songwriter, gaggle, source_notes, lines_to_remove


def clean_lyrics(raw_html, lines_to_remove=None):
    """Clean HTML content into simplified lyrics preserving musical formatting tags."""
    text = raw_html

    # Remove metadata lines that were extracted
    if lines_to_remove:
        for line in lines_to_remove:
            text = text.replace(line, "", 1)

    # Strip WP block comments
    text = re.sub(r"<!--\s*/?wp:\S+.*?-->", "", text, flags=re.DOTALL)

    # Convert <br> variants to newlines
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)

    # Convert paragraph boundaries to double newlines
    text = re.sub(r"</p>\s*<p[^>]*>", "\n\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</?p[^>]*>", "\n", text, flags=re.IGNORECASE)

    # Convert list items to newlines
    text = re.sub(r"</li>\s*<li[^>]*>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</?li[^>]*>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</?[uo]l[^>]*>", "\n", text, flags=re.IGNORECASE)

    # Strip all HTML tags EXCEPT preserved ones
    def strip_non_preserved(match):
        tag = match.group(0)
        tag_name_match = re.match(r"</?(\w+)", tag)
        if tag_name_match and tag_name_match.group(1).lower() in PRESERVE_TAGS:
            return tag
        return ""

    text = re.sub(r"<[^>]+>", strip_non_preserved, text)

    # Decode HTML entities
    text = unescape(text)
    # nbsp
    text = text.replace("\xa0", " ")

    # Trim and collapse excessive newlines
    text = text.strip()
    text = re.sub(r"\n{3,}", "\n\n", text)
    # Trim trailing whitespace on each line
    text = "\n".join(line.rstrip() for line in text.split("\n"))

    return text


def strip_all_tags(text):
    """Strip ALL HTML tags for spreadsheet plain text."""
    return re.sub(r"<[^>]+>", "", text)


def levenshtein(s1, s2):
    if len(s1) < len(s2):
        return levenshtein(s2, s1)
    if len(s2) == 0:
        return len(s1)
    prev = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        curr = [i + 1]
        for j, c2 in enumerate(s2):
            curr.append(min(curr[j] + 1, prev[j + 1] + 1, prev[j] + (c1 != c2)))
        prev = curr
    return prev[-1]


def parse_main_songs():
    tree = ET.parse(SOURCES_DIR / "main-songs.xml")
    root = tree.getroot()
    items = root.findall(".//item")

    songs = []
    idx = 0

    for item in items:
        status = item.find("wp:status", NS)
        if status is None or status.text != "publish":
            continue
        post_type = item.find("wp:post_type", NS)
        if post_type is None or post_type.text != "post":
            continue

        idx += 1
        song_id = f"main-{idx:04d}"
        wp_id = int(item.find("wp:post_id", NS).text)
        title = (item.find("title").text or "").strip()
        creator = (item.find("dc:creator", NS).text or "").strip()
        raw_content = item.find("content:encoded", NS).text or ""
        pub_date = (item.find("wp:post_date", NS).text or "")[:10]

        # Custom fields from postmeta
        meta_tune = get_meta(item, "tune")
        meta_songwriter = get_meta(item, "lyrics_by")
        meta_gaggle = get_meta(item, "gaggle")
        meta_youtube = get_meta(item, "youtube_link")
        # Check both youtube_link_ and youtube_link variations
        if not meta_youtube:
            meta_youtube = get_meta(item, "youtube_link_")
        meta_youtube_2 = get_meta(item, "youtube_link_2")
        meta_key = get_meta(item, "key_or_starting_note")
        meta_date_written = get_meta(item, "date_written_or_updated")

        # Categories and tags
        original_categories = get_categories(item)
        original_tags = get_tags(item)

        # Map categories to issues
        issues = [c for c in original_categories if c in MAIN_CATEGORIES_VALID]

        # Gaggle from creator mapping
        if creator in WEB_GRANNY_CREATORS:
            gaggle = "Unknown (Web Granny)"
        elif creator in CREATOR_TO_GAGGLE:
            gaggle = CREATOR_TO_GAGGLE[creator]
        else:
            gaggle = creator  # Use raw creator as fallback

        # Override gaggle with postmeta if available, normalizing free-text
        if meta_gaggle:
            gaggle = GAGGLE_POSTMETA_NORMALIZE.get(meta_gaggle, meta_gaggle)

        # Extract metadata from content body
        (
            content_tune,
            content_songwriter,
            content_gaggle,
            source_notes,
            lines_to_remove,
        ) = extract_metadata_from_content(raw_content)

        # Use postmeta first, fall back to content extraction
        tune = meta_tune or content_tune
        songwriter = meta_songwriter or content_songwriter

        # Resolve Web Granny gaggle from content if possible
        if gaggle == "Unknown (Web Granny)" and content_gaggle:
            gaggle = content_gaggle

        # Clean lyrics
        lyrics = clean_lyrics(raw_content, lines_to_remove if not meta_tune else [])

        needs_review = False
        review_notes = []

        if gaggle == "Unknown (Web Granny)":
            needs_review = True
            review_notes.append(
                f"Posted by Web Granny account ({creator}), gaggle could not be determined from content."
            )

        if not tune:
            needs_review = True
            review_notes.append("No tune name found in postmeta or content body.")

        song = {
            "id": song_id,
            "source": "main",
            "original_wp_id": wp_id,
            "title": title,
            "lyrics": lyrics,
            "tune": tune,
            "songwriter": songwriter,
            "gaggle": gaggle,
            "issues": issues,
            "key_or_starting_note": meta_key,
            "youtube_link": meta_youtube,
            "youtube_link_2": meta_youtube_2,
            "date_written_or_updated": meta_date_written,
            "date_published": pub_date,
            "original_creator": creator,
            "original_categories": original_categories,
            "original_tags": original_tags,
            "source_notes": source_notes,
            "duplicate_of": None,
            "needs_review": needs_review,
            "review_notes": "; ".join(review_notes) if review_notes else "",
        }
        songs.append(song)

    return songs


def parse_seattle_songs():
    tree = ET.parse(SOURCES_DIR / "seattle-songs.xml")
    root = tree.getroot()
    items = root.findall(".//item")

    songs = []
    idx = 0

    for item in items:
        status = item.find("wp:status", NS)
        if status is None or status.text != "publish":
            continue
        post_type = item.find("wp:post_type", NS)
        if post_type is None or post_type.text != "post":
            continue

        idx += 1
        song_id = f"seattle-{idx:04d}"
        wp_id = int(item.find("wp:post_id", NS).text)
        title = (item.find("title").text or "").strip()
        raw_content = item.find("content:encoded", NS).text or ""
        pub_date = (item.find("wp:post_date", NS).text or "")[:10]

        original_categories = get_categories(item)
        original_tags = get_tags(item)

        # Seattle: all songs are gaggle=Seattle
        gaggle = "Seattle"

        # Tags are songwriter names in Seattle
        songwriter = ""
        review_notes = []
        if len(original_tags) == 1:
            songwriter = original_tags[0]
        elif len(original_tags) > 1:
            songwriter = original_tags[0]
            review_notes.append(
                f"Multiple tags (potential songwriters): {', '.join(original_tags)}. Using first."
            )

        # Extract metadata from content body
        (
            content_tune,
            content_songwriter,
            content_gaggle,
            source_notes,
            lines_to_remove,
        ) = extract_metadata_from_content(raw_content)

        tune = content_tune
        if not songwriter and content_songwriter:
            songwriter = content_songwriter

        lyrics = clean_lyrics(raw_content, lines_to_remove)

        needs_review = bool(review_notes)

        song = {
            "id": song_id,
            "source": "seattle",
            "original_wp_id": wp_id,
            "title": title,
            "lyrics": lyrics,
            "tune": tune,
            "songwriter": songwriter,
            "gaggle": gaggle,
            "issues": [],  # Not mapped yet — pending approval
            "key_or_starting_note": "",
            "youtube_link": "",
            "youtube_link_2": "",
            "date_written_or_updated": "",
            "date_published": pub_date,
            "original_creator": "Jo-Hanna",
            "original_categories": original_categories,
            "original_tags": original_tags,
            "source_notes": source_notes,
            "duplicate_of": None,
            "needs_review": needs_review,
            "review_notes": "; ".join(review_notes) if review_notes else "",
        }
        songs.append(song)

    return songs


def build_seattle_category_mapping(seattle_songs):
    seattle_cats = set()
    for s in seattle_songs:
        for c in s["original_categories"]:
            seattle_cats.add(c)

    mapping = {
        "Business & Economy": "Business & Economy",
        "Children & Youth": "Human & Civil Rights",
        "Covid": "Health Care/Healthcare",
        "Crime & Punishment": "Government & Politics",
        "Education": "Education",
        "Environment & Energy": "Environment & Energy",
        "GLBT": "Human & Civil Rights",
        "George W Bush": "Government & Politics",
        "Government & Politics": "Government & Politics",
        "Grannies": None,  # meta-category, not an issue
        "Guns & Violence": "Human & Civil Rights",
        "Health": "Health Care/Healthcare",
        "Healthcare": "Health Care/Healthcare",
        "Holiday Tunes": "Holiday & Celebrations",
        "Homelessness": "Human & Civil Rights",
        "Housing": "Local Issues",
        "Human & Civil Rights": "Human & Civil Rights",
        "Immigration": "Human & Civil Rights",
        "Labor": "Labor & Worker Rights",
        "Local Issues": "Local Issues",
        "Mental Health": "Health Care/Healthcare",
        "Police Brutality": "Human & Civil Rights",
        "Poverty": "Human & Civil Rights",
        "Racism": "Human & Civil Rights",
        "Religious Issues": "Human & Civil Rights",
        "Reproductive Rights": "Women's Issues",
        "Schools": "Education",
        "Self-care": None,  # meta-category, not an issue
        "Sexism": "Women's Issues",
        "Sexuality": "Human & Civil Rights",
        "Soldiers & Veterans": "Soldiers & Veterans",
        "Trump": "Government & Politics",
        "Uncategorized": None,
        "Voting": "Government & Politics",
        "WTO": "World Issues",
        "War & Peace": "War & Peace",
        "Women": "Women's Issues",
        "older- revamp?": None,  # editorial tag, not an issue
    }

    output = {}
    for cat in sorted(seattle_cats):
        mapped = mapping.get(cat)
        output[cat] = {
            "maps_to": mapped,
            "notes": ""
            if mapped
            else "Proposed: discard (not an issue category)"
            if cat in {"Grannies", "Self-care", "Uncategorized", "older- revamp?"}
            else f"Unmapped — needs manual review",
        }

    return output


def detect_duplicates(main_songs, seattle_songs):
    main_by_title = {}
    for s in main_songs:
        key = s["title"].strip().lower()
        main_by_title.setdefault(key, []).append(s)

    duplicate_pairs = []

    for sea in seattle_songs:
        sea_title = sea["title"].strip().lower()

        # Exact match
        if sea_title in main_by_title:
            for main_match in main_by_title[sea_title]:
                sea["duplicate_of"] = main_match["id"]
                sea["needs_review"] = True
                note = f"Potential duplicate of {main_match['id']} (exact title match). Main site version from {main_match['gaggle']}, Seattle version from Seattle. Content may differ — review needed."
                if sea["review_notes"]:
                    sea["review_notes"] += "; " + note
                else:
                    sea["review_notes"] = note
                duplicate_pairs.append((main_match, sea))
            continue

        # Fuzzy match
        for main_title, main_matches in main_by_title.items():
            dist = levenshtein(sea_title, main_title)
            if dist <= 3 and dist > 0:
                main_match = main_matches[0]
                sea["duplicate_of"] = main_match["id"]
                sea["needs_review"] = True
                note = f"Potential duplicate of {main_match['id']} (fuzzy title match, distance={dist}). Main: \"{main_match['title']}\", Seattle: \"{sea['title']}\". Review needed."
                if sea["review_notes"]:
                    sea["review_notes"] += "; " + note
                else:
                    sea["review_notes"] = note
                duplicate_pairs.append((main_match, sea))
                break
            # Title contained within another
            if len(sea_title) > 5 and len(main_title) > 5:
                if sea_title in main_title or main_title in sea_title:
                    main_match = main_matches[0]
                    sea["duplicate_of"] = main_match["id"]
                    sea["needs_review"] = True
                    note = f"Potential duplicate of {main_match['id']} (title substring match). Main: \"{main_match['title']}\", Seattle: \"{sea['title']}\". Review needed."
                    if sea["review_notes"]:
                        sea["review_notes"] += "; " + note
                    else:
                        sea["review_notes"] = note
                    duplicate_pairs.append((main_match, sea))
                    break

    return duplicate_pairs


def generate_spreadsheet(all_songs, duplicate_pairs):
    wb = Workbook()

    # Main sheet
    ws = wb.active
    ws.title = "All Songs"
    headers = [
        "ID",
        "Source",
        "Title",
        "Gaggle",
        "Songwriter",
        "Tune",
        "Issues",
        "Key/Starting Note",
        "YouTube Link",
        "Date Written",
        "Lyrics",
        "Duplicate Of",
        "Needs Review",
        "Review Notes",
    ]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # Sort: needs_review TRUE first, then title
    sorted_songs = sorted(all_songs, key=lambda s: (not s["needs_review"], s["title"].lower()))

    for row_idx, song in enumerate(sorted_songs, 2):
        ws.cell(row=row_idx, column=1, value=song["id"])
        ws.cell(row=row_idx, column=2, value=song["source"])
        ws.cell(row=row_idx, column=3, value=song["title"])
        ws.cell(row=row_idx, column=4, value=song["gaggle"])
        ws.cell(row=row_idx, column=5, value=song["songwriter"])
        ws.cell(row=row_idx, column=6, value=song["tune"])
        ws.cell(row=row_idx, column=7, value=", ".join(song["issues"]))
        ws.cell(row=row_idx, column=8, value=song["key_or_starting_note"])
        ws.cell(row=row_idx, column=9, value=song["youtube_link"])
        ws.cell(row=row_idx, column=10, value=song["date_written_or_updated"])
        # Plain text lyrics for spreadsheet
        plain_lyrics = strip_all_tags(song["lyrics"])
        # Truncate to 32000 chars (Excel cell limit is 32767)
        ws.cell(row=row_idx, column=11, value=plain_lyrics[:32000])
        ws.cell(row=row_idx, column=12, value=song["duplicate_of"] or "")
        ws.cell(row=row_idx, column=13, value="TRUE" if song["needs_review"] else "FALSE")
        ws.cell(row=row_idx, column=14, value=song["review_notes"])

    # Set column widths
    col_widths = [12, 8, 40, 25, 25, 30, 30, 15, 30, 15, 50, 12, 12, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    ws.auto_filter.ref = f"A1:N{len(sorted_songs) + 1}"

    # Duplicates sheet
    ws_dup = wb.create_sheet("Duplicates")
    dup_headers = [
        "Main ID",
        "Main Title",
        "Main Gaggle",
        "Main Lyrics (first 200 chars)",
        "Seattle ID",
        "Seattle Title",
        "Seattle Gaggle",
        "Seattle Lyrics (first 200 chars)",
        "Notes",
    ]
    for col, h in enumerate(dup_headers, 1):
        cell = ws_dup.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, (main_s, sea_s) in enumerate(duplicate_pairs, 2):
        ws_dup.cell(row=row_idx, column=1, value=main_s["id"])
        ws_dup.cell(row=row_idx, column=2, value=main_s["title"])
        ws_dup.cell(row=row_idx, column=3, value=main_s["gaggle"])
        ws_dup.cell(row=row_idx, column=4, value=strip_all_tags(main_s["lyrics"])[:200])
        ws_dup.cell(row=row_idx, column=5, value=sea_s["id"])
        ws_dup.cell(row=row_idx, column=6, value=sea_s["title"])
        ws_dup.cell(row=row_idx, column=7, value=sea_s["gaggle"])
        ws_dup.cell(row=row_idx, column=8, value=strip_all_tags(sea_s["lyrics"])[:200])
        ws_dup.cell(row=row_idx, column=9, value=sea_s["review_notes"])

    # Stats sheet
    ws_stats = wb.create_sheet("Stats")
    return wb


def generate_stats(all_songs, main_songs, seattle_songs, duplicate_pairs):
    total = len(all_songs)
    main_count = len(main_songs)
    seattle_count = len(seattle_songs)

    # Gaggle distribution
    gaggle_counts = {}
    for s in all_songs:
        g = s["gaggle"]
        gaggle_counts[g] = gaggle_counts.get(g, 0) + 1

    # Issue distribution
    issue_counts = {}
    for s in all_songs:
        for iss in s["issues"]:
            issue_counts[iss] = issue_counts.get(iss, 0) + 1

    # Metadata stats
    has_tune = sum(1 for s in all_songs if s["tune"])
    has_songwriter = sum(1 for s in all_songs if s["songwriter"])
    has_youtube = sum(1 for s in all_songs if s["youtube_link"])
    has_key = sum(1 for s in all_songs if s["key_or_starting_note"])
    has_date = sum(1 for s in all_songs if s["date_written_or_updated"])

    # Quality stats
    needs_review = sum(1 for s in all_songs if s["needs_review"])
    web_granny_unresolved = sum(
        1 for s in all_songs if "Web Granny" in s["gaggle"]
    )
    no_issues = sum(1 for s in all_songs if not s["issues"])

    lines = []
    lines.append("SONG ARCHIVE CONSOLIDATION — STATS REPORT")
    lines.append("=" * 42)
    lines.append(f"Total songs parsed: {total}")
    lines.append(f"  Main site: {main_count}")
    lines.append(f"  Seattle: {seattle_count}")
    lines.append("")
    lines.append("Gaggle distribution:")
    for g, c in sorted(gaggle_counts.items(), key=lambda x: -x[1]):
        lines.append(f"  {g}: {c}")
    lines.append("")
    lines.append("Issue distribution:")
    for iss, c in sorted(issue_counts.items(), key=lambda x: -x[1]):
        lines.append(f"  {iss}: {c}")
    lines.append("")
    lines.append("Metadata extraction:")
    lines.append(f"  Songs with tune extracted: {has_tune} / {total}")
    lines.append(f"  Songs with songwriter extracted: {has_songwriter} / {total}")
    lines.append(f"  Songs with YouTube link: {has_youtube} / {total}")
    lines.append(f"  Songs with key/starting note: {has_key} / {total}")
    lines.append(f"  Songs with date written: {has_date} / {total}")
    lines.append("")
    lines.append("Data quality:")
    lines.append(f"  Songs needing review: {needs_review}")
    lines.append(f"  Duplicate pairs found: {len(duplicate_pairs)}")
    lines.append(f"  Songs from Web Granny accounts (gaggle unresolved): {web_granny_unresolved}")
    lines.append(f"  Songs with no category/issue: {no_issues}")
    lines.append("")
    seattle_unmapped = sum(
        1 for s in all_songs if s["source"] == "seattle" and not s["issues"]
    )
    seattle_has_any = any(s["issues"] for s in all_songs if s["source"] == "seattle")
    if seattle_has_any:
        if seattle_unmapped > 0:
            lines.append(
                f"Seattle category mapping: APPLIED ({seattle_unmapped} songs have no issue — original categories were all discarded)"
            )
        else:
            lines.append("Seattle category mapping: APPLIED")
    else:
        lines.append(
            "Seattle category mapping: PENDING APPROVAL (see data/seattle-category-mapping.json)"
        )

    return "\n".join(lines)


def populate_stats_sheet(wb, stats_text):
    ws = wb["Stats"]
    for row_idx, line in enumerate(stats_text.split("\n"), 1):
        ws.cell(row=row_idx, column=1, value=line)
    ws.column_dimensions["A"].width = 70


def main():
    print("Parsing main site songs...")
    main_songs = parse_main_songs()
    print(f"  Found {len(main_songs)} published songs")

    print("Parsing Seattle songs...")
    seattle_songs = parse_seattle_songs()
    print(f"  Found {len(seattle_songs)} published songs")

    print("Detecting duplicates...")
    duplicate_pairs = detect_duplicates(main_songs, seattle_songs)
    print(f"  Found {len(duplicate_pairs)} potential duplicate pairs")

    all_songs = main_songs + seattle_songs

    # Save consolidated JSON
    json_path = BASE_DIR / "songs-consolidated.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_songs, f, indent=2, ensure_ascii=False)
    print(f"Saved {json_path}")

    # Save Seattle category mapping
    cat_mapping = build_seattle_category_mapping(seattle_songs)
    cat_path = BASE_DIR / "seattle-category-mapping.json"
    with open(cat_path, "w", encoding="utf-8") as f:
        json.dump(cat_mapping, f, indent=2, ensure_ascii=False)
    print(f"Saved {cat_path}")

    # Generate stats
    stats = generate_stats(all_songs, main_songs, seattle_songs, duplicate_pairs)

    # Save stats report
    stats_path = BASE_DIR / "stats-report.txt"
    with open(stats_path, "w", encoding="utf-8") as f:
        f.write(stats)
    print(f"Saved {stats_path}")

    # Generate spreadsheet
    print("Generating review spreadsheet...")
    wb = generate_spreadsheet(all_songs, duplicate_pairs)
    populate_stats_sheet(wb, stats)
    xlsx_path = BASE_DIR / "songs-review.xlsx"
    wb.save(xlsx_path)
    print(f"Saved {xlsx_path}")

    # Print stats
    print()
    print(stats)


if __name__ == "__main__":
    main()
